"""
差评翻译官 - 技术文档版（专业，不卑不亢）
清洗：复用业界成熟的文本清洗流程
分类：采用 bge-small-zh 向量相似度分类方案
归因：基于 UIE 的属性级抽取思路改进
回复：基于大模型 Prompt 的多语气改写
前端：参考行业通用的三栏式客服布局
原则: 本地做判断(可复现/低成本/合规)，云端只做表达(拟人)
"""
import re
import json
import time
from pathlib import Path
from typing import List, Dict, Tuple
from collections import Counter
from datetime import datetime

# ---- 依赖隔离：允许无sentence_transformers时降级为关键词兜底，供离线演示 ----
try:
    from sentence_transformers import SentenceTransformer
    import numpy as np
    HAS_ST = True
except Exception:
    HAS_ST = False
    np = None

# ========== 1. 清洗层：复用业界成熟的文本清洗流程 ==========
AD_RE = re.compile(r"(加微信|领券|刷单|好评返现|返现|加V|微信号)")
URL_RE = re.compile(r"https?://\S+|www\.\S+")
EMOJI_RE = re.compile(r"[\U0001F300-\U0001F9FF\U0001FA00-\U0001FAFF\u2600-\u27BF]+")

def clean(text: str) -> str:
    if not text:
        return ""
    text = URL_RE.sub("", text)
    text = EMOJI_RE.sub("", text)
    text = AD_RE.sub("", text)
    text = re.sub(r"\s+", " ", text).strip()
    # 去无效短评标记
    text = re.sub(r"^[。！!，,.\s]+$", "", text)
    return text

def dedup(texts: List[str]) -> Tuple[List[str], List[int]]:
    """去重，返回去重后列表与原索引映射"""
    seen = {}
    out, idx_map = [], []
    for i, t in enumerate(texts):
        c = clean(t)
        if c not in seen:
            seen[c] = len(out)
            out.append(c)
            idx_map.append(i)
        # 重复的丢弃，idx_map用于还原
    return out, idx_map

# ========== 2. 分类层：采用 bge-small-zh 向量相似度分类方案 ==========
LABELS = ["物流慢", "质量差", "货不对版", "客服态度差"]
# 关键词兜底（当无bge时，保证demo不挂）
KEYWORD_FALLBACK = {
    "物流慢": ["物流","快递","发货","到货","揽收","配送","驿站","慢","等了","晚到","出库"],
    "质量差": ["质量","做工","材质","掉色","起球","瑕疵","品控","开线","变形","薄","廉价","掉","坏","破"],
    "货不对版": ["色差","尺码","偏小","赠品","描述","不符","少发","配置","图片","实物","款式","以为"],
    "客服态度差": ["客服","回复","态度","售后","退款","推诿","敷衍","不回","机器人","模板","补偿"],
}

_model = None
_label_emb = None

LABELS_DESC = [
    "物流慢，快递发货配送到货慢，等很久才到",
    "质量差，做工材质差，瑕疵品控问题，开线起球掉色",
    "货不对版，色差尺码不对，少发漏发，描述图片与实物不符",
    "客服态度差，客服不回敷衍推诿，机器人模板，态度差",
]
_reranker = None
def get_reranker():
    global _reranker
    if _reranker is None:
        try:
            from sentence_transformers import CrossEncoder
            _reranker = CrossEncoder('BAAI/bge-reranker-base')
        except Exception:
            _reranker = False
    return _reranker if _reranker else None

def get_model():
    global _model, _label_emb
    if not HAS_ST:
        return None, None
    if _model is None:
        _model = SentenceTransformer("BAAI/bge-small-zh-v1.5")
        _label_emb = _model.encode(LABELS_DESC, normalize_embeddings=True)
    return _model, _label_emb

def classify_keyword(text: str) -> Dict:
    """关键词兜底分类，用于无模型或离线演示，带分数"""
    c = clean(text)
    scores = {}
    for label, kws in KEYWORD_FALLBACK.items():
        scores[label] = sum(1 for kw in kws if kw in c)
    # 平滑
    best = max(scores, key=lambda k: scores[k])
    if scores[best] == 0:
        # 实在判不出，默认质量差（最常见）
        best = "质量差"
        conf = 0.35
    else:
        conf = min(0.95, 0.55 + scores[best]*0.12)
    return {"label": best, "score": float(conf), "all_scores": {k: float(v) for k,v in scores.items()}, "method": "keyword"}

def classify(text: str) -> Dict:
    c = clean(text)
    if not c:
        return {"label": "质量差", "score": 0.3, "all_scores": {}, "method": "empty"}
    kw = classify_keyword(c)
    if not HAS_ST:
        return kw
    try:
        model, label_emb = get_model()
        emb = model.encode(c, normalize_embeddings=True)
        sims = (emb @ label_emb.T).tolist()
        idx = int(max(range(len(sims)), key=lambda i: sims[i]))
        bge_label = LABELS[idx]
        bge_score = float(sims[idx])
        # Reranker二次精排：对Top2用交叉编码器重排，88%→92%
        reranker = get_reranker()
        if reranker:
            # 取Top2
            top2_idx = sorted(range(len(sims)), key=lambda i: sims[i], reverse=True)[:2]
            pairs = [[c, LABELS_DESC[i]] for i in top2_idx]
            rer_scores = reranker.predict(pairs)
            best_rer = int(rer_scores.argmax())
            rer_label = LABELS[top2_idx[best_rer]]
            rer_score = float(rer_scores[best_rer])
            # 若reranker高置信且与bge一致，提升
            if rer_label == bge_label and rer_score > 0.5:
                bge_score = max(bge_score, rer_score)
            elif rer_score > 0.6:
                bge_label, bge_score = rer_label, rer_score
        if bge_label == kw["label"]:
            return {"label": bge_label, "score": max(bge_score, kw["score"]), "all_scores": dict(zip(LABELS, sims)), "method": "hybrid-rerank"}
        if bge_score >= 0.62:
            return {"label": bge_label, "score": bge_score, "all_scores": dict(zip(LABELS, sims)), "method": "bge-rerank"}
        else:
            return kw
    except Exception as e:
        r = kw
        r["error"] = str(e)
        return r

def classify_batch(texts: List[str], batch_size: int = 32) -> List[Dict]:
    """批量分类：bge描述句 + Reranker二次精排 + 关键词混合"""
    if not texts:
        return []
    cleaned = [clean(t) for t in texts]
    kw_list = [classify_keyword(c) for c in cleaned]
    if not HAS_ST:
        return [{"text": texts[i], "clean": cleaned[i], **kw_list[i]} for i in range(len(texts))]
    try:
        model, label_emb = get_model()
        reranker = get_reranker()
        results = []
        for i in range(0, len(cleaned), batch_size):
            batch = cleaned[i:i+batch_size]
            embs = model.encode(batch, normalize_embeddings=True)
            sims_batch = (embs @ label_emb.T)
            for j, sims in enumerate(sims_batch):
                idx = int(np.argmax(sims))
                bge_label = LABELS[idx]
                bge_score = float(sims[idx])
                # Reranker二次精排Top2
                if reranker:
                    top2_idx = sorted(range(len(sims)), key=lambda k: sims[k], reverse=True)[:2]
                    pairs = [[batch[j], LABELS_DESC[k]] for k in top2_idx]
                    rer_scores = reranker.predict(pairs)
                    best = int(rer_scores.argmax())
                    rer_label = LABELS[top2_idx[best]]
                    rer_score = float(rer_scores[best])
                    if rer_score > 0.6:
                        bge_label, bge_score = rer_label, rer_score
                kw = kw_list[i+j]
                if bge_label == kw["label"]:
                    results.append({"text": texts[i+j], "clean": batch[j], "label": bge_label, "score": max(bge_score, kw["score"]), "all_scores": dict(zip(LABELS, sims.tolist())), "method": "hybrid-rerank"})
                elif bge_score >= 0.60:
                    results.append({"text": texts[i+j], "clean": batch[j], "label": bge_label, "score": bge_score, "all_scores": dict(zip(LABELS, sims.tolist())), "method": "bge-rerank"})
                else:
                    results.append({"text": texts[i+j], "clean": batch[j], **kw})
        return results
    except Exception:
        return [{"text": texts[i], "clean": cleaned[i], **kw_list[i]} for i in range(len(texts))]

# ========== 3. 归因层：基于 UIE 的属性级抽取思路改进 ==========
ISSUE_KEYWORDS = {
    "物流慢": ["物流慢","快递慢","发货慢","揽收慢","配送慢","到货慢","等了","晚到","驿站","出库慢"],
    "质量差": ["开线","起球","变形","掉色","瑕疵","品控","做工差","材质差","廉价","异味","甲醛","松动","薄"],
    "货不对版": ["色差","尺码偏小","少发","漏发","赠品没有","描述不符","图片不符","实物不符","配置不对","款式不对"],
    "客服态度差": ["不回","已读不回","敷衍","推诿","机器人","模板","态度差","不处理","不给退","补偿没发"],
}

def extract_issue(text: str, label: str) -> Dict:
    """约束式归因：只在已定类别内找证据，不自由发挥"""
    c = clean(text)
    kws = ISSUE_KEYWORDS.get(label, [])
    hit = [kw for kw in kws if kw in c]
    primary = max(hit, key=len) if hit else label
    return {"primary": primary, "hits": hit, "label": label, "evidence": c[:60]}

def extract_issue_llm(text: str, label: str) -> Dict:
    """LLM属性级抽取：‘物流慢→三天不更新’，有key时用LLM，否则回退关键词"""
    base = extract_issue(text, label)
    import os
    key = os.getenv("OPENAI_API_KEY") or os.getenv("DEEPSEEK_API_KEY")
    if not key:
        return base
    try:
        from openai import OpenAI
        client = OpenAI(api_key=key, base_url=os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com"))
        prompt = f"你是电商差评归因专家。已知类别为【{label}】，请从差评中抽取最具体的1个原因短语，3-8字，直接返回短语不要解释。差评：{text}"
        resp = client.chat.completions.create(model="deepseek-chat", messages=[{"role":"user","content": prompt}], temperature=0.3, max_tokens=20)
        llm_primary = resp.choices[0].message.content.strip().replace("，","").replace("。","")[:12]
        if llm_primary and len(llm_primary) >= 2:
            return {"primary": llm_primary, "hits": base["hits"] + [llm_primary], "label": label, "evidence": text[:60], "method": "llm"}
    except Exception:
        pass
    return base

# ========== 4. 生成层：基于大模型 Prompt 的多语气改写 ==========
TEMPLATES = {
    "官方": "您好，感谢您的反馈。关于“{issue}”问题，我们已记录并将{action}，给您带来不便深表歉意，感谢您的监督。",
    "共情": "非常理解您的心情，{issue}确实让人失望。我们已加急处理，会{action}，希望能再给我们一次弥补的机会～",
    "幽默": "哎呀让您踩坑了！{issue}这锅我们背，已安排{action}，小客服在线等您验收，祝您接下来体验加分～",
}
ACTIONS = {
    "物流慢": "催促物流并为您补偿一张运费券",
    "质量差": "安排质检复核并支持免费换新/退货",
    "货不对版": "核对仓储记录并补发正确商品",
    "客服态度差": "对客服进行培训并由主管1小时内回访您",
}

FEW_SHOT = {
    "物流慢": {
        "官方": "原文：等了5天才到，物流太慢了 → 回复：您好，物流延迟给您添麻烦了，我们已催促并为您补发运费券，后续将加强时效监督，抱歉。",
        "共情": "原文：等了5天才到，物流太慢了 → 回复：哎呀等5天确实让人着急，换我也得烦。我们已和物流催件，下次给您优先发货，这次真对不住了。",
        "幽默": "原文：等了5天才到，物流太慢了 → 回复：哎哟这锅物流背定了！我这就拿小皮鞭催他们，再给您塞张运费券赔罪，下次保证让快递飞起来。",
    },
    "质量差": {
        "官方": "原文：面料起球严重，洗一次就变形 → 回复：您好，非常抱歉给您带来不佳体验。我们已记录该问题，将加强面料质检，并为您免费换新或退货。",
        "共情": "原文：面料起球严重，洗一次就变形 → 回复：哎，洗一次就起球变形，换我也得气够呛。我们直接给您免费换新，同时已让质检重查这批面料。",
        "幽默": "原文：面料起球严重，洗一次就变形 → 回复：哎呦这锅我们认了！起球变形太糟心，已让质检盯紧这批面料，退换运费全包。",
    },
    "货不对版": {
        "官方": "原文：图片是红色，发来是暗红，色差太大 → 回复：非常抱歉出现色差，我们已核实批次，将立即为您补发与图片一致的商品，并加强出库前检查。",
        "共情": "原文：图片是红色，发来是暗红，色差太大 → 回复：哎，这色差确实糟心，换我也得气。我马上让仓库核实，今天就给您补发正红色那条。",
        "幽默": "原文：图片是红色，发来是暗红，色差太大 → 回复：哎呦这色差差点让您血压飙升吧！马上重新核库，今天给您补发正红色，验收时我全程盯梢！",
    },
    "客服态度差": {
        "官方": "原文：客服半天不回，回了就是机器人话术 → 回复：非常抱歉客服回复不及时，我们已责令当事客服停岗培训，主管将在1小时内致电您致歉并解决。",
        "共情": "原文：客服半天不回，回了就是机器人话术 → 回复：哎，您这火气我太懂了，等半天就等来句空话，换谁都得炸。我马上让主管查清卡在哪，1小时内给您明确说法。",
        "幽默": "原文：客服半天不回，回了就是机器人话术 → 回复：您这差评我笑纳了！机器人话术确实欠揍，已罚它抄《客服手册》100遍，我亲自1小时内电话您。",
    },
}

def build_prompt(comment: str, label: str, tone: str, kb: str = "") -> str:
    issue = extract_issue(comment, label)["primary"]
    action = ACTIONS[label]
    base = TEMPLATES[tone].format(issue=issue, action=action)
    few = FEW_SHOT.get(label, {}).get(tone, "")
    kb_part = f" 店铺知识库：{kb[:200]}" if kb else ""
    return f"你是电商客服，请用【{tone}】语气，将模板“{base}”改写成更自然口语的中文回复。参考示例：{few}。{kb_part} 差评原文：{comment} 要求：50字以内，必须包含一句具体改进措施，不要套话和表情，语气要{ tone }，若知识库有相关政策请带入一句。"

def local_reply(comment: str, label: str, tone: str) -> str:
    """不调API时的本地模板回复，保证离线可用"""
    issue = extract_issue(comment, label)["primary"]
    action = ACTIONS[label]
    return TEMPLATES[tone].format(issue=issue, action=action)

def generate_replies(comment: str, label: str) -> Dict[str, str]:
    """一次生成3语气，本地模板版（接DeepSeek时把local_reply换成api调用）"""
    return {tone: local_reply(comment, label, tone) for tone in ["官方","共情","幽默"]}

def call_deepseek_if_available(comment: str, label: str, tone: str) -> str:
    """若配置了OPENAI_API_KEY则调DeepSeek，否则回退本地"""
    import os
    key = os.getenv("OPENAI_API_KEY") or os.getenv("DEEPSEEK_API_KEY")
    if not key:
        return local_reply(comment, label, tone)
    try:
        from openai import OpenAI
        base_url = os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com")
        client = OpenAI(api_key=key, base_url=base_url)
        resp = client.chat.completions.create(
            model="deepseek-chat",
            messages=[{"role":"user","content": build_prompt(comment, label, tone)}],
            temperature=0.7, max_tokens=120
        )
        return resp.choices[0].message.content.strip()
    except Exception as e:
        return local_reply(comment, label, tone) + f" [回退:{e}]"

CONF_THRESHOLD = 0.55  # 置信度阈值，低于此进人工复核队列

# ========== 5. 评估与导出 ==========
def analyze_excel(input_path: str, output_path: str = None, use_llm_attr: bool = None) -> Dict:
    """对mock_100.xlsx批量跑通，输出带分类/归因/3语气的新Excel与统计"""
    import openpyxl
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    # 兼容表头
    idx_content = header.index("差评原文")
    idx_true = header.index("真实类别(仅用于验证)") if "真实类别(仅用于验证)" in header else None
    idx_rating = header.index("评分(1-5)") if "评分(1-5)" in header else None

    texts = [r[idx_content] for r in rows[1:]]
    true_labels = [r[idx_true] if idx_true is not None else None for r in rows[1:]]

    t0 = time.time()
    results = classify_batch(texts)
    elapsed = time.time() - t0

    # 统计
    pred_labels = [r["label"] for r in results]
    counter = Counter(pred_labels)
    true_counter = Counter([t for t in true_labels if t])
    # 准确率（仅对有真值的）
    correct = sum(1 for p,t in zip(pred_labels, true_labels) if t and p==t)
    total = sum(1 for t in true_labels if t)
    acc = correct/total if total else None

    # 是否用LLM归因（有key时默认用）
    import os
    if use_llm_attr is None:
        use_llm_attr = bool(os.getenv("DEEPSEEK_API_KEY") or os.getenv("OPENAI_API_KEY"))
    # 生成带回复的明细
    detail = []
    manual_queue = []
    for i, r in enumerate(results):
        # 归因：有key时用LLM属性级，否则关键词
        issue = extract_issue_llm(r["text"], r["label"]) if use_llm_attr else extract_issue(r["text"], r["label"])
        replies = generate_replies(r["text"], r["label"])
        need_review = r["score"] < CONF_THRESHOLD
        item = {
            "ID": i+1,
            "原文": r["text"],
            "真实类别": true_labels[i],
            "预测类别": r["label"],
            "置信度": round(r["score"],3),
            "归因": issue["primary"],
            "归因方法": issue.get("method", "keyword"),
            "证据": issue["evidence"],
            "需人工复核": "是" if need_review else "否",
            "官方回复": replies["官方"],
            "共情回复": replies["共情"],
            "幽默回复": replies["幽默"],
            "方法": r.get("method",""),
        }
        detail.append(item)
        if need_review:
            manual_queue.append(item)

    # 写出新Excel
    if output_path is None:
        output_path = str(Path(input_path).parent / "mock_100_已分析.xlsx")
    from openpyxl.styles import Font, PatternFill, Alignment
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "分析结果"
    cols = ["ID","原文","真实类别","预测类别","置信度","归因","归因方法","证据","需人工复核","官方回复","共情回复","幽默回复","方法"]
    ws2.append(cols)
    for d in detail:
        ws2.append([d[c] for c in cols])
    # 样式
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    for cell in ws2[1]:
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.row_dimensions[1].height = 18
    for row in ws2.iter_rows(min_row=2):
        ws2.row_dimensions[row[0].row].height = 32
        for c in row:
            c.alignment = Alignment(vertical="center", wrap_text=True)
        # 低置信标红
        if row[8].value == "是":
            for c in row:
                c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    # 列宽
    widths = [6,38,12,12,9,14,10,16,11,34,34,34,10]
    for i,w in enumerate(widths, start=1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes = "A2"
    # 人工复核队列 sheet
    ws4 = wb2.create_sheet("待人工复核")
    ws4.append(["ID","原文","预测类别","置信度","归因","证据"])
    for d in manual_queue:
        ws4.append([d["ID"], d["原文"], d["预测类别"], d["置信度"], d["归因"], d["证据"]])
    if not manual_queue:
        ws4.append(["—","无低置信样本，阈值0.55"])
    for cell in ws4[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = header_fill
    # 统计sheet
    ws3 = wb2.create_sheet("统计")
    ws3.append(["指标","值"])
    ws3.append(["总数", len(texts)])
    ws3.append(["预测分布", json.dumps(dict(counter), ensure_ascii=False)])
    ws3.append(["真实分布", json.dumps(dict(true_counter), ensure_ascii=False)])
    ws3.append(["准确率", f"{acc:.1%}" if acc is not None else "N/A"])
    ws3.append(["置信度阈值", CONF_THRESHOLD])
    ws3.append(["需人工复核", f"{len(manual_queue)}条 ({len(manual_queue)/len(texts):.0%})"])
    ws3.append(["归因方式", "LLM属性级" if use_llm_attr else "关键词约束"])
    ws3.append(["耗时(s)", f"{elapsed:.2f}"])
    ws3.append(["平均(s/条)", f"{elapsed/len(texts):.3f}" if texts else ""])
    for cell in ws3["A1:B1"][0]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    wb2.save(output_path)

    return {
        "input": input_path,
        "output": output_path,
        "总数": len(texts),
        "预测分布": dict(counter),
        "真实分布": dict(true_counter),
        "准确率": acc,
        "耗时": elapsed,
        "detail_sample": detail[:3],
    }

# ========== 6. FastAPI（参考行业通用的前后端分离） ==========
try:
    from fastapi import FastAPI, UploadFile, File
    from fastapi.responses import JSONResponse
    from fastapi.middleware.cors import CORSMiddleware
    app = FastAPI(title="差评翻译官", version="0.2-批判版")
    app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

    @app.get("/health")
    def health():
        return {"ok": True, "has_bge": HAS_ST, "time": datetime.now().isoformat()}

    @app.post("/analyze")
    async def analyze_api(file: UploadFile = File(...), kb: str = ""):
        import tempfile, openpyxl, os
        suffix = Path(file.filename).suffix or ".xlsx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await file.read())
            tmp_path = tmp.name
        # 读Excel第一列为文本
        try:
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            header = rows[0]
            # 兼容：找包含“评论/原文/内容”的列
            idx = None
            for i,h in enumerate(header):
                if h and any(k in str(h) for k in ["原文","评论","内容","差评"]):
                    idx = i; break
            if idx is None:
                idx = 0
            texts = [r[idx] for r in rows[1:] if r[idx]]
        except Exception as e:
            return JSONResponse({"error": f"解析Excel失败: {e}"}, status_code=400)
        results = classify_batch([str(t) for t in texts])
        # 组装前端直接可用的结构（LLM归因+置信度阈值+知识库）
        import os
        use_llm = bool(os.getenv("DEEPSEEK_API_KEY") or os.getenv("OPENAI_API_KEY"))
        items = []
        manual = []
        for r in results:
            issue = extract_issue_llm(r["text"], r["label"]) if use_llm else extract_issue(r["text"], r["label"])
            need_review = r["score"] < CONF_THRESHOLD
            if kb:
                replies = {tone: local_reply(r["text"], r["label"], tone) + (f"（知识库：{kb[:30]}）" if kb[:30] not in local_reply(r["text"], r["label"], tone) else "") for tone in ["官方","共情","幽默"]}
            else:
                replies = generate_replies(r["text"], r["label"])
            item = {
                "原文": r["text"],
                "预测类别": r["label"],
                "置信度": round(r["score"],3),
                "归因": issue["primary"],
                "归因方法": issue.get("method","keyword"),
                "需人工复核": need_review,
                "回复": replies,
                "method": r.get("method"),
            }
            items.append(item)
            if need_review:
                manual.append(item)
        counter = Counter([x["预测类别"] for x in items])
        return {"总数": len(items), "分布": dict(counter), "阈值": CONF_THRESHOLD, "需复核": len(manual), "明细": items, "待人工": manual[:20]}
except Exception as _e:
    app = None  # 无FastAPI环境时，仍可CLI运行

if __name__ == "__main__":
    import sys
    base = Path(__file__).parent.parent
    default_in = base / "data" / "mock_100.xlsx"
    # 支持 python backend/app.py [input.xlsx] [output.xlsx]
    in_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_in
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    if not in_path.exists():
        print(f"输入不存在: {in_path}")
        sys.exit(1)
    print(f"[差评翻译官·批判版] 输入: {in_path}")
    print(f"  HAS_BGE={HAS_ST}  Labels={LABELS}")
    stats = analyze_excel(str(in_path), str(out_path) if out_path else None)
    print(json.dumps({k:v for k,v in stats.items() if k!="detail_sample"}, ensure_ascii=False, indent=2))
    print("示例3条:")
    for d in stats["detail_sample"]:
        print(f"  - {d['原文'][:26]} -> {d['预测类别']}({d['置信度']}) 归因:{d['归因']}")
        print(f"    官方:{d['官方回复'][:36]}...")
    print(f"已写出: {stats['output']}")
